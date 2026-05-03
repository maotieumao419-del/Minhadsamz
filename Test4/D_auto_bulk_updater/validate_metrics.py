import os
import pandas as pd
import openpyxl
import logging

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
INPUT_1_DIR = os.path.join(BASE_DIR, "data", "input 1") # Amazon Bulk Files
INPUT_2_DIR = os.path.join(BASE_DIR, "data", "input 2") # Internal PPC Files

def read_amazon_bulk(file_path):
    """Robust reader for Amazon Bulk files with fallback engines."""
    logging.info(f"Reading Master File: {file_path}")
    try:
        # Standard Amazon sheet is "Sponsored Products Campaigns"
        xls = pd.ExcelFile(file_path)
        target_sheet = next((s for s in xls.sheet_names if "Sponsored Product" in s), None)
        if not target_sheet:
            logging.warning(f"No 'Sponsored Product' sheet found in {file_path}")
            return None
        return pd.read_excel(file_path, sheet_name=target_sheet, dtype=str)
    except Exception as e:
        logging.warning(f"Pandas default failed, trying engine='openpyxl': {e}")
        try:
            return pd.read_excel(file_path, engine='openpyxl', dtype=str)
        except Exception as e2:
            logging.error(f"Failed to read bulk file: {e2}")
            return None

def extract_campaign_blocks(df):
    """Groups data by Campaign ID and extracts SKUs, Targets, and Metrics."""
    if df is None: return {}
    df.columns = [str(c).strip() for c in df.columns]
    
    # Fill metrics with 0 if blank (User requirement)
    for col in ['Impressions', 'Clicks', 'Orders']:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype(int)
        else:
            df[col] = 0

    sku_map = {} # { SKU: { (CampName, Target): { 'MatchType', 'Bid', 'Placements', 'Metrics': (Imp, Click, Order) } } }
    
    # Group by Campaign ID
    for camp_id, group in df.groupby('Campaign ID'):
        # 1. Campaign Details
        camp_row = group[group['Entity'] == 'Campaign']
        if camp_row.empty: continue
        camp_name = camp_row['Campaign Name'].iloc[0]
        portfolio_id = camp_row.get('Portfolio ID', camp_row.get('Portfolio Id', pd.Series([""]))).iloc[0]

        # 2. Extract SKU
        skus = group[group['Entity'] == 'Product Ad']['SKU'].unique()
        if len(skus) == 0: continue

        # 4. Extract Keywords & Metrics
        kw_rows = group[group['Entity'] == 'Keyword']
        for _, r in kw_rows.iterrows():
            target = r.get('Keyword Text', '')
            metrics = (r['Impressions'], r['Clicks'], r['Orders'])
            
            for sku in skus:
                sku = str(sku).strip()
                if sku not in sku_map: sku_map[sku] = {'Portfolio ID': portfolio_id, 'Data': {}}
                sku_map[sku]['Data'][(str(camp_name).strip(), str(target).strip())] = {
                    'Metrics': metrics
                }
    return sku_map

def validate_internal_file(internal_path, master_sku_map):
    logging.info(f"Validating internal file: {internal_path}")
    try:
        wb = openpyxl.load_workbook(internal_path, data_only=True)
    except Exception as e:
        logging.error(f"Failed to load workbook: {e}")
        return

    total_checked = 0
    total_mismatch = 0

    for sheet_name in wb.sheetnames:
        if sheet_name == 'Listing':
            continue

        if sheet_name not in master_sku_map:
            continue

        ws = wb[sheet_name]
        master_data = master_sku_map[sheet_name]['Data']

        headers_row2 = [str(ws.cell(row=2, column=c).value).strip() for c in range(1, ws.max_column + 1)]
        
        try:
            camp_idx = (headers_row2.index('Campaign Name') if 'Campaign Name' in headers_row2 else 1) + 1
            target_idx = (headers_row2.index('Target') if 'Target' in headers_row2 else 3) + 1
        except ValueError:
            continue

        # Find the LAST block's Impression, Click, Order (assuming it's the rightmost one created)
        imp_idx = click_idx = order_idx = None
        for c in range(len(headers_row2), 0, -1):
            val = str(headers_row2[c-1]).strip()
            if val == 'Impression' and not imp_idx:
                imp_idx = c
                click_idx = c + 1
                order_idx = c + 2
                break
        
        if not imp_idx:
            continue

        logging.info(f" -> Checking sheet: {sheet_name} (Col {imp_idx}, {click_idx}, {order_idx})")

        for r in range(3, ws.max_row + 1):
            c_name = str(ws.cell(row=r, column=camp_idx).value or "").strip()
            t_val = str(ws.cell(row=r, column=target_idx).value or "").strip()

            if not c_name or not t_val or c_name == "None" or t_val == "None":
                continue

            key = (c_name, t_val)
            if key not in master_data:
                continue

            expected_imp, expected_click, expected_order = master_data[key]['Metrics']
            
            actual_imp = ws.cell(row=r, column=imp_idx).value
            actual_click = ws.cell(row=r, column=click_idx).value
            actual_order = ws.cell(row=r, column=order_idx).value

            # Default to 0 if None
            try: actual_imp = int(actual_imp) if actual_imp is not None else 0
            except: actual_imp = 0
            
            try: actual_click = int(actual_click) if actual_click is not None else 0
            except: actual_click = 0
            
            try: actual_order = int(actual_order) if actual_order is not None else 0
            except: actual_order = 0

            is_match = True
            mismatch_details = []
            if actual_imp != expected_imp:
                is_match = False
                mismatch_details.append(f"Impression (Expected: {expected_imp}, Actual: {actual_imp})")
            if actual_click != expected_click:
                is_match = False
                mismatch_details.append(f"Click (Expected: {expected_click}, Actual: {actual_click})")
            if actual_order != expected_order:
                is_match = False
                mismatch_details.append(f"Order (Expected: {expected_order}, Actual: {actual_order})")

            total_checked += 1
            if not is_match:
                total_mismatch += 1
                logging.error(f"Mismatch in {sheet_name} -> Row {r} | Camp: {c_name} | Target: {t_val}")
                for d in mismatch_details:
                    logging.error(f"   -> {d}")

    logging.info(f"--- VALIDATION SUMMARY for {os.path.basename(internal_path)} ---")
    logging.info(f"Total rows checked: {total_checked}")
    if total_mismatch == 0:
        logging.info("SUCCESS: All checked metrics match the Amazon Master file perfectly! 🎉")
    else:
        logging.warning(f"FAILURE: Found {total_mismatch} discrepancies. ⚠️")

def main():
    bulk_files = [f for f in os.listdir(INPUT_1_DIR) if f.startswith('BulkSheetExport') and f.endswith('.xlsx')]
    internal_files = [f for f in os.listdir(INPUT_2_DIR) if f.startswith('PPC_') and f.endswith('_UPDATED.xlsx')]
    
    if not bulk_files or not internal_files:
        logging.error("Missing BulkSheetExport (Input 1) or PPC_*_UPDATED.xlsx (Input 2) files.")
        return

    for bf in bulk_files:
        df_bulk = read_amazon_bulk(os.path.join(INPUT_1_DIR, bf))
        if df_bulk is not None:
            master_sku_map = extract_campaign_blocks(df_bulk)
            for ip in internal_files:
                validate_internal_file(os.path.join(INPUT_2_DIR, ip), master_sku_map)

if __name__ == "__main__":
    main()
