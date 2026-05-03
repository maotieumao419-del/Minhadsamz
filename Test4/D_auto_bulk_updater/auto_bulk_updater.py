import os
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment
import json
import logging
from datetime import datetime
import re
import glob

# === CONFIGURATION ===
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
INPUT_1_DIR = os.path.join(BASE_DIR, "data", "input 1")
INPUT_2_DIR = os.path.join(BASE_DIR, "data", "input 2")
OUTPUT_DIR = os.path.join(BASE_DIR, "data", "output")
RULES_FILE = "Rule_Engine.json"

AMAZON_TEMPLATE_COLUMNS = [
    'Product', 'Entity', 'Operation', 'Campaign ID', 'Ad Group ID', 'Portfolio ID',
    'Ad ID', 'Keyword ID', 'Product Targeting ID', 'Campaign Name', 'Ad Group Name',
    'Start Date', 'End Date', 'Targeting Type', 'State', 'Daily Budget', 'SKU',
    'Ad Group Default Bid', 'Bid', 'Keyword Text', 'Native Language Keyword',
    'Native Language Locale', 'Match Type', 'Bidding Strategy', 'Placement',
    'Percentage', 'Product Targeting Expression', 'Audience ID', 'Shopper Cohort Percentage',
    'Shopper Cohort Type'
]

os.makedirs(OUTPUT_DIR, exist_ok=True)

def get_real_max_row(ws, col_idx=2):
    """
    Finds the actual last row containing data by scanning upwards.
    col_idx=2 assumes Column B (Campaign Name) is a reliable column to check for text.
    """
    for row in range(ws.max_row, 0, -1):
        cell_value = ws.cell(row=row, column=col_idx).value
        if cell_value is not None and str(cell_value).strip() != "":
            return row
    return 1 # Fallback if sheet is completely empty

def load_rules():
    rule_path = os.path.join(BASE_DIR, RULES_FILE)
    if os.path.exists(rule_path):
        try:
            with open(rule_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            logging.error(f"Error reading Rule_Engine.json: {e}")
    
    # Auto-generate template if missing or failed to read
    default_rules = {
        "rules": [
            {
                "rule_name": "Bleeder_Keyword",
                "conditions": { "clicks_min": 11, "orders_max": 0 },
                "action": "Paused",
                "log_template": "Paused (Bleeder): {clicks} clicks, {orders} orders"
            }
        ]
    }
    with open(rule_path, 'w', encoding='utf-8') as f:
        json.dump(default_rules, f, indent=2)
    return default_rules

import math

def safe_id(val):
    """Safely cast Excel float IDs (e.g. 123.0) to strict string integers."""
    if pd.isna(val) or val is None: return ""
    v_str = str(val).strip()
    if v_str.lower() == "nan" or v_str == "": return ""
    try:
        return str(int(float(val)))
    except (ValueError, TypeError):
        return v_str

def normalize_str(val):
    """Bulletproof normalization handling NaNs and collapsing double spaces."""
    if val is None or pd.isna(val): return ""
    s = str(val).lower().strip()
    if s == "nan": return ""
    s = re.sub(r'\s+', ' ', s) # Collapse multiple spaces into one
    return s

def clean_match_type(val):
    """Removes brackets from Match Type for normalization (e.g. [Phrase] -> phrase)."""
    val = normalize_str(val)
    return val.replace("[", "").replace("]", "")

def get_match_type_from_ghi_chu(val):
    """Extracts Match Type from Column E (Ghi chú) which often has format '[Match Type] - ...'"""
    if not val: return ""
    match = re.search(r'\[(.*?)\]', str(val))
    return match.group(1).strip() if match else ""

def evaluate_rules(metrics, match_type, rules_config):
    """
    Evaluates rules against metrics and match_type.
    Returns: list of actions to take.
    """
    results = []
    clicks = metrics.get('Click', 0)
    orders = metrics.get('Order', 0)
    impressions = metrics.get('Impression', 0)

    for rule in rules_config.get('rules', []):
        cond = rule.get('conditions', {})
        match = True
        if 'clicks_min' in cond and clicks < cond['clicks_min']: match = False
        if 'clicks_max' in cond and clicks > cond['clicks_max']: match = False
        if 'orders_min' in cond and orders < cond['orders_min']: match = False
        if 'orders_max' in cond and orders > cond['orders_max']: match = False
        if 'impressions_min' in cond and impressions < cond['impressions_min']: match = False
        if 'impressions_max' in cond and impressions > cond['impressions_max']: match = False

        if match:
            # Check if there is match-type specific logic
            mt_logic = rule.get('match_type_logic', {})
            if mt_logic:
                # Find specific logic for this match type (case-insensitive)
                logic = None
                for key in mt_logic:
                    if key.lower() == str(match_type).lower():
                        logic = mt_logic[key]
                        break
                
                if logic:
                    log_msg = rule.get('log_template', "").format(
                        clicks=clicks, orders=orders, impressions=impressions, match_type=match_type
                    )
                    res = logic.copy()
                    res['log_msg'] = log_msg
                    results.append(res)
            else:
                # Standard action
                log_msg = rule.get('log_template', "").format(
                    clicks=clicks, orders=orders, impressions=impressions, match_type=match_type
                )
                results.append({
                    'action': rule.get('action'),
                    'log_msg': log_msg,
                    'target_entity': 'Keyword' # Default
                })
    
    return results

def main():
    rules_config = load_rules()
    current_date = datetime.now().strftime("%d%m%Y")
    
    # 1. Discover Files
    master_files = glob.glob(os.path.join(INPUT_1_DIR, "BulkSheetExport_*.xlsx"))
    internal_files = glob.glob(os.path.join(INPUT_2_DIR, "PPC_*_UPDATED.xlsx"))

    if not internal_files:
        logging.error(f"No UPDATED internal files found in: {INPUT_2_DIR}")
        return
    
    internal_path = internal_files[0] # Take the first one found
    logging.info(f"Discovered Internal File: {internal_path}")

    # 2. Build Dynamic ID Mapping Engine
    campaign_map = {} # Key: Normalized Campaign Name
    keyword_map = {}  # Key: (Norm. Campaign Name, Norm. Keyword Text, Norm. Match Type)
    
    if master_files:
        master_path = master_files[0]
        logging.info(f"Building ID Map from Master File: {master_path}")
        try:
            df_master = pd.read_excel(master_path, sheet_name='Sponsored Products Campaigns')
            
            # Dynamic column fallback
            kw_col = 'Keyword Text' if 'Keyword Text' in df_master.columns else 'Keyword'
            pt_col = 'Product Targeting Expression' if 'Product Targeting Expression' in df_master.columns else 'Product Targeting'
            
            for _, row in df_master.iterrows():
                entity = normalize_str(row.get('Entity'))
                n_camp = normalize_str(row.get('Campaign Name'))
                n_mt = normalize_str(row.get('Match Type'))
                
                # Smart Target Extraction (Keyword or ASIN)
                raw_kw = row.get(kw_col)
                if pd.isna(raw_kw) or str(raw_kw).strip() == "":
                    raw_kw = row.get(pt_col)
                n_kw = normalize_str(raw_kw)
                
                c_id = safe_id(row.get('Campaign ID'))
                ag_id = safe_id(row.get('Ad Group ID'))
                kw_id = safe_id(row.get('Keyword ID'))
                ag_name = str(row.get('Ad Group Name', '')).strip() if pd.notna(row.get('Ad Group Name')) else ''
                
                raw_bid = row.get('Bid')
                try:
                    current_bid = float(raw_bid) if pd.notna(raw_bid) and str(raw_bid).strip() != "" else 0.0
                except (ValueError, TypeError):
                    current_bid = 0.0

                if n_camp and c_id:
                    campaign_map[n_camp] = c_id
                
                # ONLY map actual Keywords/Targets to prevent pollution
                if entity in ['keyword', 'product targeting'] and n_camp and (n_kw or n_mt):
                    keyword_map[(n_camp, n_kw, n_mt)] = {
                        'Campaign ID': c_id,
                        'Ad Group ID': ag_id,
                        'Keyword ID': kw_id,
                        'Ad Group Name': ag_name,
                        'Current_Bid': current_bid
                    }
                    
            logging.info(f"Built ID Map: {len(campaign_map)} Campaigns, {len(keyword_map)} Keywords/Targets.")
        except Exception as e:
            logging.error(f"Error building Dynamic ID map: {e}")

    # 3. Process Internal File
    wb_internal = openpyxl.load_workbook(internal_path)
    triggered_actions = [] # List of actions for Bulk and Internal Logging

    for sheet_name in wb_internal.sheetnames:
        if sheet_name == 'Listing': continue
        ws = wb_internal[sheet_name]
        logging.info(f"Processing sheet: {sheet_name}")

        # Locate Latest Metrics Block (Row 2)
        headers_row2 = [str(ws.cell(row=2, column=c).value).strip() for c in range(1, ws.max_column + 1)]
        note_cols = [i+1 for i, h in enumerate(headers_row2) if h == 'Note']
        if not note_cols:
            logging.warning(f"No 'Note' column found in Row 2 of sheet {sheet_name}")
            continue
        
        last_note_idx = note_cols[-1]
        order_idx = last_note_idx - 1
        click_idx = last_note_idx - 2
        imp_idx = last_note_idx - 3

        # Fixed Header Indices
        try:
            camp_col = headers_row2.index('Campaign Name') + 1
            target_col = headers_row2.index('Target') + 1
            ghi_chu_col = 5 # As per requirement "Column E/5"
        except ValueError:
            logging.warning(f"Missing essential headers in {sheet_name}. Skipping.")
            continue

        def safe_int(val):
            if val is None or str(val).strip() == "" or pd.isna(val):
                return 0
            try:
                return int(float(val))
            except (ValueError, TypeError):
                return 0

        for r in range(3, ws.max_row + 1):
            campaign_name = ws.cell(row=r, column=camp_col).value
            target_text = ws.cell(row=r, column=target_col).value
            if not campaign_name or not target_text: continue

            ghi_chu_val = ws.cell(row=r, column=ghi_chu_col).value
            match_type = get_match_type_from_ghi_chu(ghi_chu_val)

            metrics = {
                'Impression': safe_int(ws.cell(row=r, column=imp_idx).value),
                'Click': safe_int(ws.cell(row=r, column=click_idx).value),
                'Order': safe_int(ws.cell(row=r, column=order_idx).value)
            }

            actions = evaluate_rules(metrics, match_type, rules_config)
            
            if actions:
                for act in actions:
                    # Perform Reverse Lookup
                    ids = {}
                    n_camp = normalize_str(campaign_name)
                    
                    if act.get('target_entity') == 'Bidding Adjustment':
                        ids['Campaign ID'] = campaign_map.get(n_camp, "")
                    else:
                        # Keyword lookup requires normalization and match type cleaning
                        n_kw = normalize_str(target_text)
                        n_mt = clean_match_type(match_type)
                        ids = keyword_map.get((n_camp, n_kw, n_mt), {})
                        if not ids:
                            logging.warning(f"ID Lookup Failed: Camp='{n_camp}', Target='{n_kw}', Match='{n_mt}'")
                    
                    triggered_actions.append({
                        'SKU': sheet_name,
                        'Row': r,
                        'Campaign Name': campaign_name,
                        'Target': target_text,
                        'Match Type': match_type,
                        'Action_Data': act,
                        'NoteCol': last_note_idx,
                        'IDs': ids
                    })

    # 4. Generate Amazon Upload File
    if triggered_actions:
        logging.info(f"Generating Amazon Bulk Upload file for {len(triggered_actions)} actions...")
        upload_rows = []
        
        for item in triggered_actions:
            act = item['Action_Data']
            ids = item.get('IDs', {})
            
            # Initialize row with empty strings for all standard columns
            row_dict = {col: "" for col in AMAZON_TEMPLATE_COLUMNS}
            
            row_dict['Product'] = "Sponsored Products"
            
            target_entity = act.get('target_entity', 'Keyword')
            row_dict['Entity'] = target_entity
            
            action_val = str(act.get('action', 'Update')).strip()
            if action_val.lower() in ['paused', 'enabled', 'archived']:
                row_dict['Operation'] = 'Update'
                row_dict['State'] = action_val.title()
            else:
                row_dict['Operation'] = 'Update'
            row_dict['Campaign Name'] = item['Campaign Name']
            
            # Fill IDs and related names from Master lookup
            row_dict['Campaign ID'] = ids.get('Campaign ID', '')
            row_dict['Ad Group ID'] = ids.get('Ad Group ID', '')
            row_dict['Keyword ID'] = ids.get('Keyword ID', '')
            row_dict['Ad Group Name'] = ids.get('Ad Group Name', '')
            
            if target_entity.lower() == 'bidding adjustment':
                row_dict['Placement'] = act.get('placement_type', '')
                row_dict['Percentage'] = act.get('set_percentage', '')
            else:
                if target_entity.lower() == 'product targeting':
                    row_dict['Product Targeting Expression'] = item['Target']
                else:
                    row_dict['Keyword Text'] = item['Target']
                    
                row_dict['Match Type'] = item['Match Type']
                
                if 'adjust_bid_by' in act:
                    base_bid = ids.get('Current_Bid', 0.0)
                    try:
                        adjust_val = float(act['adjust_bid_by'])
                    except (ValueError, TypeError):
                        adjust_val = 0.0
                    new_bid = round(base_bid + adjust_val, 2)
                    row_dict['Bid'] = new_bid

                # State is already set above if applicable
            
            upload_rows.append(row_dict)

        if upload_rows:
            df_out = pd.DataFrame(upload_rows, columns=AMAZON_TEMPLATE_COLUMNS)
            bulk_out = os.path.join(OUTPUT_DIR, f"Amazon_Upload_Ready_{current_date}.xlsx")
            
            try:
                with pd.ExcelWriter(bulk_out, engine='xlsxwriter') as writer:
                    df_out.to_excel(writer, index=False, sheet_name='Sponsored Products Campaigns')
                    workbook  = writer.book
                    worksheet = writer.sheets['Sponsored Products Campaigns']
                    text_fmt  = workbook.add_format({'num_format': '@'}) # Force Text format
                    
                    # Apply Text format and fixed width to all columns
                    for col_num in range(len(AMAZON_TEMPLATE_COLUMNS)):
                        worksheet.set_column(col_num, col_num, 18, text_fmt)
                        
                logging.info(f"Amazon Bulk Upload file generated successfully: {bulk_out}")
            except Exception as e:
                logging.error(f"Error generating Amazon upload file: {e}")

    # 5. Update Internal File (Logging)
    logging.info("Updating Internal File with logs...")
    for item in triggered_actions:
        ws = wb_internal[item['SKU']]
        log_text = item['Action_Data']['log_msg']
        cell = ws.cell(row=item['Row'], column=item['NoteCol'], value=log_text)
        
        # Highlight Red if Paused
        if item['Action_Data'].get('action') == 'Paused':
            cell.font = Font(color="FF0000", bold=True)
        
    # Using a timestamped name to avoid PermissionError if file is open
    timestamp = datetime.now().strftime("%H%M%S")
    internal_out = os.path.join(OUTPUT_DIR, f"Internal_File_Optimized_{current_date}_{timestamp}.xlsx")
    wb_internal.save(internal_out)
    logging.info(f"Optimized Internal File saved: {internal_out}")

if __name__ == "__main__":
    main()

