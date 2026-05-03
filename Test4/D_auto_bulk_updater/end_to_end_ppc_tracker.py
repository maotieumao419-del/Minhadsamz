import os
import pandas as pd
import openpyxl
from openpyxl.styles import Border, Side, Font, Alignment
import logging
import re
from datetime import datetime
import sys
import copy

# === CONFIGURATION ===
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
INPUT_1_DIR = os.path.join(BASE_DIR, "data", "input 1") # Amazon Bulk Files
INPUT_2_DIR = os.path.join(BASE_DIR, "data", "input 2") # Internal PPC Files

def get_file_date(filename):
    """Extracts date range from filename like 'BulkSheetExport_17-20.xlsx'."""
    match = re.search(r'_(\d+-\d+)', filename)
    return match.group(1) if match else datetime.now().strftime("%d")

def read_amazon_bulk(file_path):
    """Robust reader for Amazon Bulk files with fallback engines."""
    logging.info(f"Reading Master File: {file_path}")
    try:
        # Standard Amazon sheet is "Sponsored Products Campaigns"
        xls = pd.ExcelFile(file_path)
        target_sheet = next((s for s in xls.sheet_names if "Sponsored Product" in s), None)
        if not target_sheet:
            logging.warning(f"No 'Sponsored Product' sheet found in {file_path}")
            return None
        return pd.read_excel(file_path, sheet_name=target_sheet, dtype=str)
    except Exception as e:
        logging.warning(f"Pandas default failed, trying engine='openpyxl': {e}")
        try:
            return pd.read_excel(file_path, engine='openpyxl', dtype=str)
        except Exception as e2:
            logging.error(f"Failed to read bulk file: {e2}")
            return None

def extract_campaign_blocks(df):
    """Groups data by Campaign ID and extracts SKUs, Targets, and Metrics."""
    if df is None: return {}
    df.columns = [str(c).strip() for c in df.columns]
    
    # Fill metrics with 0 if blank (User requirement)
    for col in ['Impressions', 'Clicks', 'Orders']:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype(int)
        else:
            df[col] = 0

    sku_map = {} # { SKU: { (CampName, Target): { 'MatchType', 'Bid', 'Placements', 'Metrics': (Imp, Click, Order) } } }
    
    # Group by Campaign ID
    for camp_id, group in df.groupby('Campaign ID'):
        # 1. Campaign Details
        camp_row = group[group['Entity'] == 'Campaign']
        if camp_row.empty: continue
        camp_name = camp_row['Campaign Name'].iloc[0]
        portfolio_id = camp_row.get('Portfolio ID', camp_row.get('Portfolio Id', pd.Series([""]))).iloc[0]

        # 2. Extract SKU
        skus = group[group['Entity'] == 'Product Ad']['SKU'].unique()
        if len(skus) == 0: continue

        # 3. Extract Placements
        adj_rows = group[group['Entity'] == 'Bidding Adjustment']
        placements = {'Top': '0%', 'PP': '0%'}
        for _, r in adj_rows.iterrows():
            p_type = str(r.get('Placement Type', '')).lower()
            perc = r.get('Percentage', '0')
            if 'top' in p_type: placements['Top'] = f"{perc}%"
            elif 'product page' in p_type: placements['PP'] = f"{perc}%"

        # 4. Extract Keywords & Targeting & Metrics
        kw_rows = group[group['Entity'].str.strip().str.lower().isin(['keyword', 'product targeting'])]
        for _, r in kw_rows.iterrows():
            entity_type = str(r['Entity']).strip().lower()
            if entity_type == 'keyword':
                target = r.get('Keyword Text', '')
            elif entity_type == 'product targeting':
                target = r.get('Product Targeting Expression', '')
            else:
                target = ''
            match_type = r.get('Match Type', '')
            bid = r.get('Bid', '')
            metrics = (r['Impressions'], r['Clicks'], r['Orders'])
            
            for sku in skus:
                sku = str(sku).strip()
                if sku not in sku_map: sku_map[sku] = {'Portfolio ID': portfolio_id, 'Data': {}}
                sku_map[sku]['Data'][(str(camp_name).strip(), str(target).strip())] = {
                    'MatchType': match_type,
                    'Bid': bid,
                    'Placements': placements,
                    'Metrics': metrics
                }
    return sku_map

def copy_style(source_cell, target_cell):
    """
    Safely copies formatting from a source cell to a target cell by extracting the StyleProxy objects.
    """
    if source_cell.has_style:
        target_cell.font = copy.copy(source_cell.font)
        target_cell.border = copy.copy(source_cell.border)
        target_cell.fill = copy.copy(source_cell.fill)
        target_cell.number_format = copy.copy(source_cell.number_format)
        target_cell.protection = copy.copy(source_cell.protection)
        target_cell.alignment = copy.copy(source_cell.alignment)

def get_real_max_row(ws, col_idx=2):
    """
    Finds the actual last row containing data by scanning upwards.
    col_idx=2 assumes Column B (Campaign Name) is a reliable column to check for text.
    """
    for row in range(ws.max_row, 0, -1):
        cell_value = ws.cell(row=row, column=col_idx).value
        if cell_value is not None and str(cell_value).strip() != "":
            return row
    return 1 # Fallback if sheet is completely empty

def process_internal_file(internal_path, master_sku_map, date_range):
    """Phase 2: 4-Column Block Sync & Metrics Injection."""
    logging.info(f" -> [DEBUG] Loading workbook: {internal_path}")
    try:
        wb = openpyxl.load_workbook(internal_path)
    except Exception as e:
        logging.error(f"Failed to load workbook: {e}")
        return
    logging.info(" -> [DEBUG] Workbook loaded successfully.")
    
    # Border style for the block
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    # 1. Update Listing (Standard Logic)
    if 'Listing' in wb.sheetnames:
        ws_l = wb['Listing']
        headers = [str(ws_l.cell(row=2, column=c).value).strip() for c in range(1, ws_l.max_column + 1)]
        sku_idx = headers.index('SKU') + 1 if 'SKU' in headers else 2
        p_id_idx = next((i+1 for i, h in enumerate(headers) if 'portfolio' in h.lower()), 6)
        
        existing_skus = {str(ws_l.cell(row=r, column=sku_idx).value).strip() for r in range(3, ws_l.max_row + 1)}
        for sku, info in master_sku_map.items():
            if sku not in existing_skus:
                new_row = get_real_max_row(ws_l, col_idx=sku_idx) + 1
                ws_l.cell(row=new_row, column=sku_idx, value=sku)
                ws_l.cell(row=new_row, column=p_id_idx, value=info['Portfolio ID'])
                if new_row > 3:
                    for c in range(1, ws_l.max_column + 1):
                        copy_style(ws_l.cell(row=new_row-1, column=c), ws_l.cell(row=new_row, column=c))
                existing_skus.add(sku)

    # 2. Update SKU Sheets
    logging.info(f" -> [DEBUG] Iterating through {len(master_sku_map)} SKUs from Master Map...")
    for sku, info in master_sku_map.items():
        if sku not in wb.sheetnames:
            logging.info(f" -> [DEBUG] Creating new sheet for SKU: {sku}")
            ws = wb.create_sheet(sku)
            # Row 1: SKU Name (Protected)
            ws.cell(row=1, column=1, value=sku)
            ws.cell(row=1, column=1).font = Font(bold=True, size=14)
            # Row 2: Headers
            headers_list = ['STT', 'Campaign Name', 'Loại Campaign', 'Target', 'Initial Note']
            for c, h in enumerate(headers_list, 1):
                ws.cell(row=2, column=c, value=h)
                ws.cell(row=2, column=c).font = Font(bold=True)
        else:
            ws = wb[sku]
        
        # TASK A: IDENTIFY OR INITIALIZE HEADERS

        # 0. GLOBAL SCORCHED EARTH FOR ROW 2 
        # Find and destroy ANY merged cells that touch Row 2 across the entire sheet
        for merged_range in list(ws.merged_cells.ranges):
            if merged_range.min_row <= 2 <= merged_range.max_row:
                try:
                    ws.unmerge_cells(str(merged_range))
                except Exception:
                    pass

        # 1. VALIDATE AND INITIALIZE BASE STRUCTURE
        logging.info(f" -> [DEBUG] Validating base structure for {sku}. Sheet max_row: {ws.max_row}")
        base_headers = ["STT", "Campaign Name", "Loại Campaign", "Target", "Note", "Trạng thái"]
        note_col_idx = None
        status_col_idx = None

        # Scan Row 2 for columns
        for col in range(1, ws.max_column + 2):
            val = str(ws.cell(row=2, column=col).value).strip().lower()
            if val in ["note", "ghi chú", "initial note"]:
                note_col_idx = col
            if val in ["trạng thái", "status"]:
                status_col_idx = col
            if note_col_idx and status_col_idx:
                break

        # If 'Note' is missing or sheet is empty, strictly build the base structure in Columns 1 to 5
        if not note_col_idx:
            logging.info(f" -> [DEBUG] Base structure missing for {sku}. Initializing columns 1-5.")
            for i, header_name in enumerate(base_headers, start=1):
                cell = ws.cell(row=2, column=i)
                cell.value = header_name
                cell.font = Font(bold=True)
                cell.border = thin_border
            note_col_idx = 5 # 'Note' is now at column 5 (E)

        # 2. CALCULATE NEXT START COLUMN
        # The metrics block must ALWAYS start AFTER the Note column.
        next_start_col = note_col_idx + 1
        while ws.cell(row=2, column=next_start_col).value is not None:
            next_start_col += 1
        logging.info(f" -> [DEBUG] Next metrics block will start at column: {next_start_col}")
        
        # 1. SCORCHED EARTH UNMERGE: Find and destroy any existing merges in the target 4 columns
        for merged_range in list(ws.merged_cells.ranges):
            # Check if the merged range intersects with our target 4 columns
            if merged_range.min_col <= next_start_col + 3 and merged_range.max_col >= next_start_col:
                try:
                    ws.unmerge_cells(str(merged_range))
                except Exception:
                    pass

        # 2. Now it is 100% safe to merge Row 1 for the Date string
        ws.merge_cells(start_row=1, start_column=next_start_col, end_row=1, end_column=next_start_col + 3)
        cell_merged = ws.cell(row=1, column=next_start_col, value=date_range)
        cell_merged.alignment = Alignment(horizontal='center')
        cell_merged.font = Font(bold=True)

        # 3. Safely write headers to Row 2
        block_headers = ['Impression', 'Click', 'Order', 'Note']
        for i, h in enumerate(block_headers):
            target_col = next_start_col + i
            # Since all merges in this column are gone, it will be a standard Cell, not a MergedCell
            target_cell = ws.cell(row=2, column=target_col)
            target_cell.value = h
            target_cell.font = Font(bold=True)
            target_cell.border = thin_border

        # Map static column indices
        headers_row2 = [str(ws.cell(row=2, column=c).value).strip() for c in range(1, next_start_col)]
        camp_idx = (headers_row2.index('Campaign Name') if 'Campaign Name' in headers_row2 else 1) + 1
        target_idx = (headers_row2.index('Target') if 'Target' in headers_row2 else 3) + 1
        stt_idx = (headers_row2.index('STT') if 'STT' in headers_row2 else 0) + 1

        # TASK B: Reverse Cleanup (Standard Logic) - Stop at Row 2
        logging.info(f" -> [DEBUG] Starting TASK B (Cleanup) for {sku}. Checking {ws.max_row} rows.")
        master_data = info['Data']
        for r in range(ws.max_row, 2, -1):
            if r <= 2: continue
            c_name = str(ws.cell(row=r, column=camp_idx).value or "").strip()
            t_val = str(ws.cell(row=r, column=target_idx).value or "").strip()
            
            # Guard Clause 1: Skip Headers (already handled by range, but safety check)
            if c_name == "Campaign Name" or t_val == "Target":
                continue

            # Guard Clause 2: PROTECT DRAFT / PRE-LAUNCH CAMPAIGNS
            if "AD_READY" in c_name.upper() or "CHƯA TẠO" in c_name.upper():
                continue
                
            # Guard Clause 3: Check Status column
            if status_col_idx:
                status_val = str(ws.cell(row=r, column=status_col_idx).value or "").strip().upper()
                if any(x in status_val for x in ["CÔNG VIỆC MỚI", "DRAFT", "TIẾP NHẬN"]):
                    continue

            if (c_name, t_val) not in master_data and c_name and t_val:
                ws.delete_rows(r)

        # TASK C: INJECT METRICS & APPEND NEW TARGETS
        logging.info(f" -> [DEBUG] Starting TASK C (Injection) for {sku}.")
        existing_rows = {} # { (Camp, Target): [row_idx1, ...] }
        for r in range(3, ws.max_row + 1):
            c_name = str(ws.cell(row=r, column=camp_idx).value).strip()
            t_val = str(ws.cell(row=r, column=target_idx).value).strip()
            if c_name and t_val and c_name != "None":
                key = (c_name, t_val)
                if key not in existing_rows:
                    existing_rows[key] = []
                existing_rows[key].append(r)

        for key, data in master_data.items(): # key is (camp_name, target_val)
            c_name, t_val = key
            metrics = data['Metrics']
            
            # 1. IF NEW TARGET -> CREATE COMPLETE ROW FIRST
            if key not in existing_rows:
                new_row = get_real_max_row(ws, col_idx=camp_idx) + 1
                
                # Explicitly write the base columns
                ws.cell(row=new_row, column=stt_idx, value=new_row - 2)
                ws.cell(row=new_row, column=camp_idx, value=c_name)
                ws.cell(row=new_row, column=3, value="Keyword (Từ khóa)") # Loại Campaign
                ws.cell(row=new_row, column=target_idx, value=t_val)
                
                # Construct Note info
                note_str = f"[{data.get('MatchType', 'N/A')}] - [{data.get('Bid', 'N/A')}] - [Top: {data['Placements'].get('Top', '0%')} / PP: {data['Placements'].get('PP', '0%')}]"
                ws.cell(row=new_row, column=note_col_idx, value=note_str)
                
                # Copy styling from the row above
                if new_row > 3:
                    for col_idx in range(1, note_col_idx + 1):
                        copy_style(ws.cell(row=new_row-1, column=col_idx), ws.cell(row=new_row, column=col_idx))
                
                existing_rows[key] = [new_row]

            # 2. INJECT METRICS (For both existing and newly created rows)
            for r_idx in existing_rows[key]:
                # Write to the 4-column block (next_start_col is the Impression column)
                ws.cell(row=r_idx, column=next_start_col, value=metrics[0])     # Impression
                ws.cell(row=r_idx, column=next_start_col + 1, value=metrics[1]) # Click
                ws.cell(row=r_idx, column=next_start_col + 2, value=metrics[2]) # Order
                
                # Apply borders to the injected cells (Imp, Click, Order, AND Note)
                for i in range(4):
                    ws.cell(row=r_idx, column=next_start_col + i).border = thin_border

    output_path = internal_path.replace(".xlsx", "_UPDATED.xlsx")
    wb.save(output_path)
    logging.info(f"Saved: {output_path}")

def main():
    # 1. Discover Files
    bulk_files = [f for f in os.listdir(INPUT_1_DIR) if f.startswith('BulkSheetExport') and f.endswith('.xlsx')]
    internal_files = [f for f in os.listdir(INPUT_2_DIR) if f.startswith('PPC_') and f.endswith('.xlsx') and "_UPDATED" not in f]
    
    if not bulk_files or not internal_files:
        logging.error("Missing BulkSheetExport (Input 1) or PPC_* (Input 2) files.")
        return

    # Assuming latest bulk file for now or processing all
    for bf in bulk_files:
        date_range = get_file_date(bf)
        df_bulk = read_amazon_bulk(os.path.join(INPUT_1_DIR, bf))
        if df_bulk is not None:
            master_sku_map = extract_campaign_blocks(df_bulk)
            # Process each internal file
            for ip in internal_files:
                process_internal_file(os.path.join(INPUT_2_DIR, ip), master_sku_map, date_range)

if __name__ == "__main__":
    main()
