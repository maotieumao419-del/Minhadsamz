import os
import pandas as pd
import openpyxl
from openpyxl.styles import Border, Side, Font, Alignment
import logging
import re
import sys
from copy import copy

# === CONFIGURATION ===
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
INPUT_DIR = os.path.join(BASE_DIR, "data", "input 1")
OUTPUT_DIR = os.path.join(BASE_DIR, "data", "input 2")

AMAZON_PREFIX = "BulkSheetExport_"
INTERNAL_PREFIX = "PPC_"

os.makedirs(INPUT_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)

def clean_sku(val):
    """Remove leading/trailing whitespace and newlines but preserve internal ones."""
    if val is None: return ""
    return str(val).strip('\r\n\t ')

def read_amazon_bulk(file_path):
    """
    Robust reader for Amazon Bulk files (XLSM/XLSX/CSV fallback).
    Tries multiple engines and encodings to handle corrupted or malformed files.
    """
    filename = os.path.basename(file_path)
    logging.info(f"Attempting to read Amazon Bulk (MASTER): {filename}")
    
    # 1. Try standard Excel engines first
    for engine in ['openpyxl', None]:
        try:
            xls = pd.ExcelFile(file_path, engine=engine)
            sheet_names = xls.sheet_names
            # Standard Amazon Bulk can have variations like "Sponsored Products Campaigns"
            target_sheet = next((s for s in sheet_names if "Sponsored Product" in s or "Sheet1" in s), None)
            
            if not target_sheet:
                logging.warning(f"Could not find 'Sponsored Product' sheet in {filename}. Available: {sheet_names}")
                continue
                
            df = pd.read_excel(file_path, sheet_name=target_sheet, engine=engine, dtype=str)
            logging.info(f"Successfully read via Excel engine: {engine or 'default'}")
            return df
        except Exception as e:
            logging.debug(f"Excel engine {engine or 'default'} failed for {filename}: {e}")

    # 2. Try CSV fallback with multiple encodings (Handles cases where Excel is corrupted but data is text-like)
    for encoding in ['utf-8', 'latin1', 'cp1252']:
        try:
            logging.info(f"Trying CSV fallback for {filename} with encoding: {encoding}")
            df = pd.read_csv(file_path, sep=None, engine='python', on_bad_lines='skip', dtype=str, encoding=encoding)
            logging.info(f"Successfully read via CSV fallback ({encoding})")
            return df
        except Exception as e:
            logging.debug(f"CSV fallback with {encoding} failed: {e}")

    logging.error(f"Critical Error: Failed to read file {filename} after all attempts.")
    return None

def extract_campaign_blocks(df):
    """
    Processes the DataFrame using the 'Campaign Block' rule.
    Returns a dictionary structured by SKU.
    """
    if df is None: return {}
    
    # Standardize column names (strip whitespace)
    df.columns = [str(c).strip() for c in df.columns]
    
    # Group by Campaign ID (or fallback to Campaign Name if ID is missing)
    camp_id_col = next((c for c in df.columns if "Campaign ID" in str(c)), "Campaign ID")
    if camp_id_col not in df.columns:
        camp_id_col = next((c for c in df.columns if "Campaign Name" in str(c)), None)
        if not camp_id_col:
            logging.error("Required column 'Campaign ID' or 'Campaign Name' not found in Bulk file!")
            return {}

    logging.info(f"Grouping bulk data by '{camp_id_col}'...")
    grouped = df.groupby(camp_id_col)
    
    sku_map = {} 
    
    for camp_id, group in grouped:
        entity_col = next((c for c in df.columns if "Entity" in str(c)), "Entity")
        if entity_col not in df.columns: continue

        # 1. Extract Core Info from 'Campaign' entity
        campaign_row = group[group[entity_col].astype(str).str.lower() == 'campaign']
        if campaign_row.empty: continue
        
        p_id_col = next((c for c in df.columns if "Portfolio ID" in str(c)), "Portfolio ID")
        p_name_col = next((c for c in df.columns if "Portfolio Name" in str(c)), "Portfolio Name")
        c_name_col = next((c for c in df.columns if "Campaign Name" in str(c)), "Campaign Name")

        portfolio_id = str(campaign_row[p_id_col].iloc[0]).strip() if p_id_col in campaign_row.columns else ""
        portfolio_name = str(campaign_row[p_name_col].iloc[0]).strip() if p_name_col in campaign_row.columns else ""
        campaign_name = str(campaign_row[c_name_col].iloc[0]).strip() if c_name_col in campaign_row.columns else f"Unknown_{camp_id}"
        
        # 2. Extract SKUs from 'Product Ad' entity
        ad_rows = group[group[entity_col].astype(str).str.lower() == 'product ad']
        sku_col = next((c for c in df.columns if "SKU" in str(c)), "SKU")
        skus = [clean_sku(s) for s in ad_rows[sku_col].unique() if clean_sku(s) and str(s).lower() != 'nan']
        
        if not skus: continue 
        
        # 3. Extract Keywords
        kw_rows = group[group[entity_col].astype(str).str.lower() == 'keyword']
        keywords = []
        kw_text_col = next((c for c in df.columns if "Keyword Text" in str(c)), "Keyword Text")
        mt_col = next((c for c in df.columns if "Match Type" in str(c)), "Match Type")
        bid_col = next((c for c in df.columns if "Bid" in str(c)), "Bid")

        for _, r in kw_rows.iterrows():
            keywords.append({
                'Text': str(r.get(kw_text_col, '')).strip(),
                'MatchType': str(r.get(mt_col, '')).strip(),
                'Bid': str(r.get(bid_col, '')).strip()
            })
            
        # 4. Extract Placements
        adj_rows = group[group[entity_col].astype(str).str.lower() == 'bidding adjustment']
        placements = {'Top': '0%', 'PP': '0%'}
        p_type_col = next((c for c in df.columns if "Placement" in str(c)), "Placement Type")
        percent_col = next((c for c in df.columns if "Percentage" in str(c)), "Percentage")

        for _, r in adj_rows.iterrows():
            p_type = str(r.get(p_type_col, '')).lower()
            percentage = str(r.get(percent_col, '0')).strip()
            if 'top' in p_type:
                placements['Top'] = f"{percentage}%"
            elif 'product page' in p_type:
                placements['PP'] = f"{percentage}%"
                
        # Map data to every SKU linked to this campaign
        for sku in skus:
            if sku not in sku_map:
                sku_map[sku] = {
                    'Portfolio ID': portfolio_id, 
                    'Portfolio Name': portfolio_name,
                    'Campaigns': {}
                }
            
            if campaign_name not in sku_map[sku]['Campaigns']:
                sku_map[sku]['Campaigns'][campaign_name] = {
                    'Keywords': keywords,
                    'Placements': placements
                }
            else:
                # In case of duplicate campaign names across bulk files
                existing_kw_texts = {kw['Text'] for kw in sku_map[sku]['Campaigns'][campaign_name]['Keywords']}
                for kw in keywords:
                    if kw['Text'] not in existing_kw_texts:
                        sku_map[sku]['Campaigns'][campaign_name]['Keywords'].append(kw)

    return sku_map

def copy_style(source_cell, target_cell):
    """Safely copy cell formatting to avoid StyleProxy errors."""
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = source_cell.number_format
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)

def get_real_max_row(ws, col_idx=2):
    """
    Finds the actual last row containing data by scanning upwards.
    col_idx=2 assumes Column B (Campaign Name) is a reliable column to check for text.
    """
    for row in range(ws.max_row, 0, -1):
        cell_value = ws.cell(row=row, column=col_idx).value
        if cell_value is not None and str(cell_value).strip() != "":
            return row
    return 1 # Fallback if sheet is completely empty

def update_listing_sheet(wb, sku_map):
    """
    Syncs the main 'Listing' sheet. 
    Adds new SKUs and fills Portfolio details.
    """
    if 'Listing' not in wb.sheetnames:
        logging.error("Sheet 'Listing' not found in internal file!")
        return
    
    ws = wb['Listing']
    header = [str(ws.cell(row=1, column=c).value).strip() for c in range(1, ws.max_column + 1)]
    
    try:
        sku_idx = (header.index('SKU') if 'SKU' in header else 1) + 1
    except ValueError:
        sku_idx = 2 # Guessing column B
        
    p_id_idx = next((i for i, h in enumerate(header) if 'portfolio id' in h.lower()), 5) + 1
    p_name_idx = next((i for i, h in enumerate(header) if 'portfolio name' in h.lower()), p_id_idx) + 1

    # Identify existing SKUs (cleaned)
    existing_skus = set()
    for r in range(2, ws.max_row + 1):
        s = clean_sku(ws.cell(row=r, column=sku_idx).value)
        if s and s != 'None':
            existing_skus.add(s)
        
    logging.info(f"Syncing Listing sheet: {len(existing_skus)} existing entries found.")
    
    # Robustly find the last data row to avoid appending after empty phantom rows
    current_last_row = get_real_max_row(ws, sku_idx)
    
    added_count = 0
    for sku, info in sku_map.items():
        if sku not in existing_skus:
            current_last_row += 1
            logging.info(f" -> Adding NEW SKU: {sku} at row {current_last_row}")
            # Add Full SKU
            ws.cell(row=current_last_row, column=sku_idx, value=sku)
            # Add Portfolio Details
            ws.cell(row=current_last_row, column=p_id_idx, value=info['Portfolio ID'])
            if p_name_idx <= ws.max_column:
                ws.cell(row=current_last_row, column=p_name_idx, value=info['Portfolio Name'])
            
            # Apply formatting from row above
            if current_last_row > 2:
                for c in range(1, ws.max_column + 1):
                    copy_style(ws.cell(row=current_last_row-1, column=c), ws.cell(row=current_last_row, column=c))
            
            existing_skus.add(sku)
            added_count += 1
            
    logging.info(f"Listing Update: Added {added_count} new SKUs.")

def update_sku_sheets(wb, sku_map):
    """
    Processes individual SKU sheets.
    Handles Excel's 31-char limit for tab names.
    """
    for sku, info in sku_map.items():
        # EXCEL LIMIT: Tab names must be <= 31 chars
        tab_name = sku[:31]
        
        # Avoid creating empty sheets if no campaigns found
        if not info['Campaigns']:
            continue

        if tab_name not in wb.sheetnames:
            ws = wb.create_sheet(tab_name)
            # Add Full SKU identifier at the top for user reference
            ws.cell(row=1, column=1, value=f"Full SKU: {sku}")
            ws.cell(row=1, column=1).font = Font(bold=True, size=12, color="0000FF")
            
            headers = ['STT', 'Campaign Name', 'Loại Campaign', 'Target', 'Ghi chú']
            for i, h in enumerate(headers):
                ws.cell(row=2, column=i+1, value=h)
                ws.cell(row=2, column=i+1).font = Font(bold=True)
                ws.cell(row=2, column=i+1).border = Border(bottom=Side(style='thin'))
            start_row = 3
        else:
            ws = wb[tab_name]
            # Detect if we have the "Full SKU" header to determine data start
            start_row = 2
            if "Full SKU:" in str(ws.cell(row=1, column=1).value):
                start_row = 3
            
        header = [str(ws.cell(row=start_row-1, column=c).value).strip() for c in range(1, ws.max_column + 1)]
        try:
            stt_idx = (header.index('STT') if 'STT' in header else 0) + 1
            camp_idx = (header.index('Campaign Name') if 'Campaign Name' in header else 1) + 1
            target_idx = (header.index('Target') if 'Target' in header else 3) + 1
            note_idx = 5
            for i, h in enumerate(header):
                if h.lower() in ['note', 'ghi chú']:
                    note_idx = i + 1
                    break
        except Exception:
            stt_idx, camp_idx, target_idx, note_idx = 1, 2, 4, 5

        # Map existing keywords in the sheet to avoid duplicates
        existing_data = {} 
        for r in range(start_row, ws.max_row + 1):
            c_name = str(ws.cell(row=r, column=camp_idx).value).strip()
            t_val = str(ws.cell(row=r, column=target_idx).value).strip()
            n_val = str(ws.cell(row=r, column=note_idx).value).strip()
            if c_name and t_val:
                existing_data[(c_name, t_val)] = n_val

        for camp_name, details in info['Campaigns'].items():
            keywords = details['Keywords']
            placements = details['Placements']
            note_val_base = f"[Top: {placements['Top']} / PP: {placements['PP']}]"
            
            for kw in keywords:
                text = str(kw['Text']).strip()
                m_type = str(kw['MatchType']).strip()
                bid = str(kw['Bid']).strip()
                note_content = f"[{m_type}] - [{bid}] - {note_val_base}"
                
                key = (camp_name, text)
                
                if key in existing_data:
                    # Update ONLY if the note is empty or Nan
                    if not existing_data[key] or existing_data[key].lower() in ['none', 'nan', '']:
                        for r in range(start_row, ws.max_row + 1):
                            match_camp = str(ws.cell(row=r, column=camp_idx).value).strip()
                            match_target = str(ws.cell(row=r, column=target_idx).value).strip()
                            if match_camp == key[0] and match_target == key[1]:
                                ws.cell(row=r, column=note_idx, value=note_content)
                                break
                else:
                    # Append new target row
                    target_row = get_real_max_row(ws, col_idx=camp_idx) + 1
                    
                    # Calculate new STT
                    last_stt = ws.cell(row=target_row - 1, column=stt_idx).value
                    try:
                        new_stt = int(last_stt) + 1 if last_stt and str(last_stt).isdigit() else 1
                    except:
                        new_stt = 1
                    
                    row_data = [new_stt, camp_name, "Keyword (Từ khóa)", text, note_content]
                    indices = [stt_idx, camp_idx, camp_idx+1, target_idx, note_idx]
                    
                    for idx, val in zip(indices, row_data):
                        ws.cell(row=target_row, column=idx, value=val)
                        
                    # Copy Style from row above
                    if target_row > start_row:
                        for c in range(1, ws.max_column + 1):
                            copy_style(ws.cell(row=target_row-1, column=c), ws.cell(row=target_row, column=c))

def main():
    print("--- AUTOMATED AMAZON BULK SYNC SYSTEM ---")
    
    # 1. Identify Source Files
    amazon_files = [f for f in os.listdir(INPUT_DIR) if f.startswith(AMAZON_PREFIX) and f.endswith('.xlsx')]
    if not amazon_files:
        logging.error(f"No Amazon Bulk files found starting with '{AMAZON_PREFIX}'")
        return

    # 2. Identify Internal Target Files
    internal_files = [f for f in os.listdir(INPUT_DIR) if f.startswith(INTERNAL_PREFIX) and f.endswith('.xlsx')]
    if not internal_files:
        logging.error(f"No internal PPC files found starting with '{INTERNAL_PREFIX}'")
        return

    logging.info(f"Detecting Workflow: {len(amazon_files)} Amazon sources -> {len(internal_files)} Internal targets.")

    # 3. Aggregate Brain (Data extracted from all Amazon files)
    master_sku_map = {}
    for af in amazon_files:
        df_bulk = read_amazon_bulk(os.path.join(INPUT_DIR, af))
        if df_bulk is not None:
            file_map = extract_campaign_blocks(df_bulk)
            # Merge logic
            for sku, info in file_map.items():
                if sku not in master_sku_map:
                    master_sku_map[sku] = info
                else:
                    for camp, details in info['Campaigns'].items():
                        master_sku_map[sku]['Campaigns'][camp] = details
    
    if not master_sku_map:
        logging.warning("No campaign-SKU data could be extracted. Check bulk file format.")
        return

    # 4. Synchronize each Internal File
    for inf in internal_files:
        path = os.path.join(INPUT_DIR, inf)
        try:
            logging.info(f"Starting Synchronization for Store File: {inf}")
            wb = openpyxl.load_workbook(path)
            
            # Update Sheets
            update_listing_sheet(wb, master_sku_map)
            update_sku_sheets(wb, master_sku_map)
            
            # Save results (Preserve original filename in output)
            save_path = os.path.join(OUTPUT_DIR, inf)
            logging.info(f"Saving synchronized file: {os.path.basename(save_path)}")
            wb.save(save_path)
            
        except Exception as e:
            logging.error(f"Critical synchronization error for {inf}: {e}")
            import traceback
            traceback.print_exc()
            
    logging.info("SYSTEM SYNC COMPLETED SUCCESSFULLY.")

if __name__ == "__main__":
    main()
